from flask import Flask, render_template, request, jsonify, send_from_directory, redirect, url_for, flash, session
from flask_sqlalchemy import SQLAlchemy
from flask_cors import CORS
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from flask_migrate import Migrate
from flask_marshmallow import Marshmallow
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from flask_wtf.csrf import generate_csrf
from flask_wtf import CSRFProtect
import os
import json
import requests
from sqlalchemy import func, case
from decimal import Decimal, ROUND_HALF_UP
from datetime import datetime, timedelta, timezone
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from config import config

# --- App and Database Setup ---
base_dir = os.path.abspath(os.path.dirname(__file__))

# Create Flask app
app = Flask(__name__)

# Load configuration
env = os.environ.get('FLASK_ENV', 'development')
app.config.from_object(config[env])

# Initialize extensions
db = SQLAlchemy(app)
migrate = Migrate(app, db)
ma = Marshmallow(app)
csrf = CSRFProtect(app)
login_manager = LoginManager(app)
login_manager.login_view = 'login'

# Configure CORS with specific origins
cors_origins = app.config.get('CORS_ORIGINS', ['*'])
CORS(app, resources={r"/api/*": {"origins": cors_origins}}, supports_credentials=True)

# Rate limiting
limiter = Limiter(
    app=app,
    key_func=get_remote_address,
    storage_uri=app.config.get('RATELIMIT_STORAGE_URL', 'memory://'),
    default_limits=["1000 per day", "200 per hour"],  # More generous defaults
    enabled=app.config.get('RATELIMIT_ENABLED', True)
)

# Exempt API routes from CSRF (they use authentication instead)
@app.before_request
def exempt_api_csrf():
    # Exempt JSON/API endpoints from CSRF checks (they use other auth).
    if request.path.startswith('/api/'):
        try:
            # More reliable way to exempt API endpoints
            csrf._exempt_views.add(request.endpoint)
            # Also set a flag that can be checked by CSRF protection
            setattr(request, 'csrf_exempt', True)
            app.logger.debug(f"CSRF exempted API endpoint: {request.endpoint}")
        except Exception as e:
            # If the private attribute access fails, log it but continue
            app.logger.debug(f"CSRF exemption failed: {e}")
            pass
    else:
        # Log non-API requests for debugging
        if request.method == 'POST' and request.path == '/login':
            app.logger.info(f"Login request - Path: {request.path}, Method: {request.method}")
            app.logger.info(f"Request headers: {dict(request.headers)}")
            app.logger.info(f"Form keys: {list(request.form.keys()) if request.form else 'No form data'}")
    
# Error Handlers
@app.errorhandler(404)
def not_found(e):
    if request.path.startswith('/api/'):
        return jsonify({'error': 'Resource not found'}), 404
    return render_template('base.html'), 404

@app.errorhandler(500)
def internal_error(e):
    db.session.rollback()
    if request.path.startswith('/api/'):
        return jsonify({'error': 'Internal server error'}), 500
    flash('An unexpected error occurred. Please try again.', 'danger')
    return redirect(url_for('dispatcher_dashboard'))

@app.errorhandler(429)
def ratelimit_handler(e):
    if request.path.startswith('/api/'):
        return jsonify({'error': 'Rate limit exceeded. Please try again later.'}), 429
    flash('Too many requests. Please slow down.', 'warning')
    return redirect(request.referrer or url_for('dispatcher_dashboard'))

# --- Language Translations ---
translations = {}
with open('translations.json', 'r', encoding='utf-8') as f:
    translations = json.load(f)

@app.context_processor
def inject_gettext():
    def _(key):
        lang = get_locale()
        return translations.get(lang, translations['en']).get(key, key)
    return dict(_=_, translations=translations)

def get_locale():
    # First check session, then cookie for persistence
    lang = session.get('language')
    if not lang:
        lang = request.cookies.get('language_preference', 'en')
        if lang:
            session['language'] = lang
    return lang or 'en'

@app.route('/change_language/<lang>')
def change_language(lang):
    # Validate language code
    if lang not in ['en', 'am', 'ti']:
        lang = 'en'  # Default to English if invalid
    
    session['language'] = lang
    
    # Also store in localStorage via cookie for persistence
    response = redirect(request.referrer or url_for('passenger_home'))
    response.set_cookie('language_preference', lang, max_age=365*24*60*60)  # 1 year
    
    return response

@app.route('/api/get_language')
def get_language():
    """API endpoint to get current language preference"""
    lang = get_locale()
    return jsonify({'language': lang})

@app.route('/api/debug/status')
def debug_status():
    """Debug endpoint to check system status"""
    try:
        # Check database connection
        db_status = "OK"
        try:
            db.session.execute(db.text("SELECT 1")).fetchone()
        except Exception as e:
            db_status = f"Error: {str(e)}"
        
        # Check tables exist
        tables = {}
        try:
            tables['admin_count'] = Admin.query.count()
            tables['setting_count'] = Setting.query.count()
            tables['driver_count'] = Driver.query.count()
            tables['passenger_count'] = Passenger.query.count()
        except Exception as e:
            tables['error'] = str(e)
        
        return jsonify({
            'status': 'running',
            'database': db_status,
            'tables': tables,
            'config': {
                'debug': app.config.get('DEBUG'),
                'upload_folder': app.config.get('UPLOAD_FOLDER')
            }
        })
    except Exception as e:
        return jsonify({'error': str(e), 'status': 'error'}), 500

@app.route('/api/debug/test-post', methods=['POST'])
@csrf.exempt
def test_post():
    """Simple POST test endpoint to debug request handling"""
    try:
        data = {
            'method': request.method,
            'content_type': request.content_type,
            'headers': dict(request.headers),
            'json_data': request.get_json(silent=True),
            'form_data': dict(request.form),
            'raw_data': request.get_data(as_text=True)[:200]  # First 200 chars
        }
        return jsonify({'success': True, 'request_info': data})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/debug/admin-users')
def debug_admin_users():
    """Debug endpoint to check admin users"""
    try:
        admins = Admin.query.all()
        admin_info = [{
            'id': admin.id,
            'username': admin.username,
            'has_password': bool(admin.password_hash)
        } for admin in admins]
        
        return jsonify({
            'total_admins': len(admins),
            'admins': admin_info
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/debug/create-test-admin', methods=['POST'])
@csrf.exempt
def create_test_admin():
    """Debug endpoint to create a test admin user"""
    try:
        # Check if test admin already exists
        existing = Admin.query.filter_by(username='admin').first()
        if existing:
            return jsonify({'message': 'Test admin already exists', 'username': 'admin'})
        
        # Create test admin
        test_admin = Admin(username='admin')
        test_admin.set_password('admin123')
        db.session.add(test_admin)
        db.session.commit()
        
        return jsonify({
            'message': 'Test admin created successfully',
            'username': 'admin',
            'password': 'admin123'
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/set_language', methods=['POST'])
@csrf.exempt
def set_language():
    """API endpoint to set language preference"""
    data = request.get_json()
    lang = data.get('language', 'en')
    
    # Validate language code
    if lang not in ['en', 'am', 'ti']:
        lang = 'en'
    
    session['language'] = lang
    return jsonify({'success': True, 'language': lang})

# --- Helper Functions ---
def to_eat(utc_dt):
    """Converts a UTC datetime object to East Africa Time (EAT)."""
    if utc_dt is None:
        return None
    return utc_dt.replace(tzinfo=timezone.utc).astimezone(timezone(timedelta(hours=3)))

# --- Custom Decorators for Role-Based Access ---
def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or session.get('user_type') != 'admin':
            # Return JSON for API requests, HTML redirect for page requests
            if request.path.startswith('/api/'):
                return jsonify({'error': 'Authentication required. Please log in as a dispatcher.'}), 401
            flash('You must be logged in as a dispatcher to view this page.', 'danger')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def passenger_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or session.get('user_type') != 'passenger':
            # Return JSON for API requests, HTML redirect for page requests
            if request.path.startswith('/api/'):
                return jsonify({'error': 'Authentication required. Please log in as a passenger.'}), 401
            flash('Please log in to access this page.', 'warning')
            return redirect(url_for('passenger_login'))
        return f(*args, **kwargs)
    return decorated_function

# --- Database Models ---
class Admin(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password_hash = db.Column(db.String(200), nullable=False)
    profile_picture = db.Column(db.String(255), nullable=True, default='static/img/default_user.svg')

    def set_password(self, password):
        self.password_hash = generate_password_hash(password)

    def check_password(self, password):
        return check_password_hash(self.password_hash, password)

class Passenger(UserMixin, db.Model):
    __tablename__ = 'passenger'
    id = db.Column(db.Integer, primary_key=True)
    passenger_uid = db.Column(db.String(20), unique=True, nullable=True, index=True)
    username = db.Column(db.String(80), nullable=False)
    phone_number = db.Column(db.String(20), unique=True, nullable=False, index=True)
    password_hash = db.Column(db.String(200), nullable=False)
    profile_picture = db.Column(db.String(255), nullable=True, default='static/img/default_user.svg')
    join_date = db.Column(db.DateTime, server_default=db.func.now())

    def set_password(self, password):
        self.password_hash = generate_password_hash(password)

    def check_password(self, password):
        return check_password_hash(self.password_hash, password)


class Driver(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    driver_uid = db.Column(db.String(20), unique=True, nullable=True, index=True)  # User-friendly ID
    name = db.Column(db.String(100), nullable=False)
    phone_number = db.Column(db.String(20), nullable=False, index=True)
    vehicle_type = db.Column(db.String(50), nullable=False, default='Bajaj', index=True)
    vehicle_details = db.Column(db.String(150), nullable=False)
    vehicle_plate_number = db.Column(db.String(50), nullable=True)
    license_info = db.Column(db.String(100), nullable=True)
    status = db.Column(db.String(20), default='Offline', nullable=False, index=True)
    profile_picture = db.Column(db.String(255), nullable=True, default='static/img/default_user.svg')
    license_document = db.Column(db.String(255), nullable=True)
    vehicle_document = db.Column(db.String(255), nullable=True)
    join_date = db.Column(db.DateTime, server_default=db.func.now())
    current_lat = db.Column(db.Float, nullable=True)
    current_lon = db.Column(db.Float, nullable=True)


class Ride(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    passenger_id = db.Column(db.Integer, db.ForeignKey('passenger.id'), nullable=False, index=True)
    driver_id = db.Column(db.Integer, db.ForeignKey('driver.id'), nullable=True, index=True)
    pickup_address = db.Column(db.String(255), nullable=True)
    pickup_lat = db.Column(db.Float, nullable=False)
    pickup_lon = db.Column(db.Float, nullable=False)
    dest_address = db.Column(db.String(255), nullable=False)
    dest_lat = db.Column(db.Float, nullable=True)
    dest_lon = db.Column(db.Float, nullable=True)
    distance_km = db.Column(db.Numeric(10, 2), nullable=False)
    fare = db.Column(db.Numeric(10, 2), nullable=False)  # Changed from Float to Numeric for money
    vehicle_type = db.Column(db.String(50), nullable=False, default='Bajaj', index=True)
    status = db.Column(db.String(20), default='Requested', nullable=False, index=True)
    request_time = db.Column(db.DateTime, server_default=db.func.now(), index=True)
    assigned_time = db.Column(db.DateTime, nullable=True)
    note = db.Column(db.String(255), nullable=True)
    payment_method = db.Column(db.String(20), nullable=False, default='Cash')

    passenger = db.relationship('Passenger', backref=db.backref('rides', lazy=True))
    driver = db.relationship('Driver', backref=db.backref('rides', lazy=True))
    feedback = db.relationship('Feedback', backref='ride', uselist=False, cascade="all, delete-orphan")


class Feedback(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    ride_id = db.Column(db.Integer, db.ForeignKey('ride.id'), unique=True, nullable=False)
    rating = db.Column(db.Integer, nullable=True)
    comment = db.Column(db.String(500), nullable=True)
    feedback_type = db.Column(db.String(50), nullable=False, default='Rating')
    details = db.Column(db.Text, nullable=True)
    is_resolved = db.Column(db.Boolean, default=False)
    submitted_at = db.Column(db.DateTime, server_default=db.func.now(timezone.utc))

class Setting(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    key = db.Column(db.String(100), unique=True, nullable=False)
    value = db.Column(db.String(255), nullable=False)

# --- Marshmallow Schemas ---
class PassengerSchema(ma.SQLAlchemyAutoSchema):
    class Meta:
        model = Passenger
        fields = ("username", "phone_number") # Only expose what's needed

class DriverSchema(ma.SQLAlchemyAutoSchema):
    avg_rating = ma.Float()
    class Meta:
        model = Driver

class RideSchema(ma.SQLAlchemyAutoSchema):
    passenger = ma.Nested(PassengerSchema)
    request_time = ma.Function(lambda obj: to_eat(obj.request_time).strftime('%I:%M %p'))
    # Rename for clarity
    user_name = ma.String(attribute="passenger.username")
    user_phone = ma.String(attribute="passenger.phone_number")

    class Meta:
        model = Ride
        # Explicitly list fields to match original output
        fields = ("id", "user_name", "user_phone", "pickup_address", "pickup_lat",
                  "pickup_lon", "dest_address", "dest_lat", "dest_lon", "fare",
                  "vehicle_type", "note", "request_time")

# Schemas for multiple items
rides_schema = RideSchema(many=True)
drivers_schema = DriverSchema(many=True)

# --- Service Layer Initialization ---
# Import services after models are defined
from services import (
    PassengerService, DriverService, RideService, 
    FeedbackService, SettingService, AdminService, AnalyticsService
)

# Initialize service instances
passenger_service = PassengerService(db, Passenger)
driver_service = DriverService(db, Driver)
ride_service = RideService(db, Ride)
feedback_service = FeedbackService(db, Feedback)
setting_service = SettingService(db, Setting)
admin_service = AdminService(db, Admin)
analytics_service = AnalyticsService(db, Ride, Driver, Passenger, Feedback)


def get_setting(key, default=None):
    """Wrapper for backward compatibility - uses setting_service"""
    return setting_service.get_setting(key, default)

def _handle_file_upload(file_storage, existing_path=None):
    """Handle file upload with security checks"""
    if not file_storage or not file_storage.filename:
        return existing_path
    
    filename = secure_filename(file_storage.filename)
    if not filename:
        return existing_path
    
    # Check file extension using Config class method
    from config import Config
    if not Config.allowed_file(filename):
        raise ValueError(f"File type not allowed. Allowed types: {app.config['ALLOWED_EXTENSIONS']}")
    
    # Create upload directory if it doesn't exist
    upload_folder = app.config.get('UPLOAD_FOLDER', os.path.join(base_dir, 'static', 'uploads'))
    os.makedirs(upload_folder, exist_ok=True)
    
    # Generate unique filename if file already exists
    name, ext = os.path.splitext(filename)
    save_path = os.path.join(upload_folder, filename)
    if os.path.exists(save_path):
        timestamp = int(datetime.now(timezone.utc).timestamp())
        filename = f"{name}_{timestamp}{ext}"
        save_path = os.path.join(upload_folder, filename)
    
    # Save file
    file_storage.save(save_path)
    
    # Delete old file if it exists and is not the default
    if existing_path and 'default_' not in existing_path:
        old_file_path = os.path.join(base_dir, existing_path)
        if os.path.exists(old_file_path):
            try:
                os.remove(old_file_path)
            except OSError:
                pass  # Ignore errors if file can't be deleted
    
    rel_path = os.path.join('static', 'uploads', filename).replace('\\', '/')
    return rel_path

# --- Authentication ---
@login_manager.user_loader
def load_user(user_id):
    user_type = session.get('user_type')
    if user_type == 'admin':
        return Admin.query.get(int(user_id))
    elif user_type == 'passenger':
        return Passenger.query.get(int(user_id))
    return None

# -- Admin Auth ---
@app.route('/login', methods=['GET', 'POST'])
@limiter.limit("5 per minute")
def login():
    if current_user.is_authenticated and session.get('user_type') == 'admin':
        return redirect(url_for('dispatcher_dashboard'))
    if request.method == 'POST':
        try:
            # Debug logging for login attempts
            app.logger.info(f"Login POST request received")
            app.logger.info(f"Content-Type: {request.content_type}")
            app.logger.info(f"Form data keys: {list(request.form.keys())}")
            app.logger.info(f"Has CSRF token: {'csrf_token' in request.form}")
            
            username = request.form.get('username', '').strip()
            password = request.form.get('password', '')
            
            app.logger.info(f"Username provided: {bool(username)}")
            app.logger.info(f"Password provided: {bool(password)}")
            
            if not username or not password:
                app.logger.warning("Missing username or password")
                flash('Username and password are required', 'danger')
                return render_template('login.html')
            
            admin = Admin.query.filter_by(username=username).first()
            if admin and admin.check_password(password):
                app.logger.info(f"Login successful for user: {username}")
                login_user(admin)
                session['user_type'] = 'admin'
                return redirect(url_for('dispatcher_dashboard'))
            else:
                app.logger.warning(f"Login failed for user: {username}")
                flash('Invalid username or password', 'danger')
                return render_template('login.html')
        except Exception as e:
            app.logger.error(f"Login error: {str(e)}")
            flash('Login error occurred', 'danger')
            return render_template('login.html'), 500
    return render_template('login.html')

# -- Passenger Auth ---
@app.route('/passenger/signup', methods=['GET', 'POST'])
@limiter.limit("3 per hour")
def passenger_signup():
    if current_user.is_authenticated and session.get('user_type') == 'passenger':
        return redirect(url_for('passenger_app'))
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        phone_number_input = request.form.get('phone_number', '').strip()
        password = request.form.get('password', '')
        
        # Validation
        if not username or not phone_number_input or not password:
            flash('All fields are required.', 'danger')
            return render_template('passenger_signup.html')
        
        if len(password) < 6:
            flash('Password must be at least 6 characters long.', 'danger')
            return render_template('passenger_signup.html')
        
        # Sanitize phone number
        phone_number = "+251" + phone_number_input
        if not phone_number_input.isdigit() or len(phone_number_input) != 9:
            flash('Invalid phone number format. Please enter 9 digits.', 'danger')
            return render_template('passenger_signup.html')
        
        # Use service layer to check if passenger exists
        existing_passenger = passenger_service.get_by_phone(phone_number)
        if existing_passenger:
            flash('Phone number already registered.', 'danger')
            return redirect(url_for('passenger_signup'))
        
        # Use service layer to create new passenger
        new_passenger = passenger_service.create(
            username=username,
            phone_number=phone_number
        )
        new_passenger.set_password(password)
        db.session.flush()
        new_passenger.passenger_uid = f"PAX-{new_passenger.id:05d}"
        passenger_service.commit()
        
        flash('Account created successfully! Please log in.', 'success')
        return redirect(url_for('passenger_login'))
    return render_template('passenger_signup.html')


@app.route('/passenger/login', methods=['GET', 'POST'])
@limiter.limit("5 per minute")
def passenger_login():
    if current_user.is_authenticated and session.get('user_type') == 'passenger':
        return redirect(url_for('passenger_app'))
    if request.method == 'POST':
        phone_number_input = request.form.get('phone_number', '').strip()
        password = request.form.get('password', '')
        
        if not phone_number_input or not password:
            flash('Phone number and password are required.', 'danger')
            return render_template('passenger_login.html')
        
        phone_number = "+251" + phone_number_input
        passenger = Passenger.query.filter_by(phone_number=phone_number).first()
        
        if passenger and passenger.check_password(password):
            login_user(passenger)
            session['user_type'] = 'passenger'
            return redirect(url_for('passenger_app'))
        else:
            flash('Invalid phone number or password.', 'danger')
    return render_template('passenger_login.html')


@app.route('/logout')
@login_required
def logout():
    user_type = session.get('user_type')
    logout_user()
    session.pop('user_type', None)
    if user_type == 'admin':
        return redirect(url_for('login'))
    else:
        return redirect(url_for('passenger_login'))

# --- Frontend Routes ---
@app.route('/')
@admin_required
def dispatcher_dashboard():
    return render_template('dashboard.html')

@app.route('/passenger')
def passenger_home():
    if current_user.is_authenticated and session.get('user_type') == 'passenger':
        return redirect(url_for('passenger_app'))
    return redirect(url_for('passenger_login'))

@app.route('/request')
@passenger_required
def passenger_app():
    return render_template('passenger.html')

@app.route('/passenger/profile', methods=['GET', 'POST'])
@passenger_required
def passenger_profile():
    passenger = current_user
    if request.method == 'POST':
        current_password = request.form.get('current_password')
        if not passenger.check_password(current_password):
            flash('Your current password is not correct.', 'danger')
            return redirect(url_for('passenger_profile'))
        
        passenger.username = request.form.get('username', passenger.username)
        
        new_password = request.form.get('new_password')
        if new_password:
            passenger.set_password(new_password)
        
        if 'profile_picture' in request.files:
            passenger.profile_picture = _handle_file_upload(request.files['profile_picture'], passenger.profile_picture)

        db.session.commit()
        flash('Profile updated successfully!', 'success')
        return redirect(url_for('passenger_profile'))

    return render_template('passenger_profile.html')


@app.route('/passenger/history')
@passenger_required
def passenger_history():
    return render_template('passenger_history.html')

@app.route('/passenger/support')
@passenger_required
def passenger_support():
    return render_template('Passenger Support.html')


@app.route('/uploads/<filename>')
def uploaded_file(filename):
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename)

# --- API Routes ---
@app.route('/api/ride-request', methods=['POST'])
@csrf.exempt
@passenger_required
@limiter.limit("10 per hour")
def request_ride():
    data = request.json
    if not data:
        return jsonify({'error': 'No data provided'}), 400
    
    # Validate required fields
    required_fields = ['pickup_lat', 'pickup_lon', 'dest_address', 'dest_lat', 'dest_lon', 'distance_km', 'fare']
    for field in required_fields:
        if field not in data or data[field] is None:
            return jsonify({'error': f'Missing required field: {field}'}), 400
    
    # Validate coordinates
    try:
        pickup_lat = float(data['pickup_lat'])
        pickup_lon = float(data['pickup_lon'])
        dest_lat = float(data['dest_lat'])
        dest_lon = float(data['dest_lon'])
        
        if not (-90 <= pickup_lat <= 90) or not (-180 <= pickup_lon <= 180):
            return jsonify({'error': 'Invalid pickup coordinates'}), 400
        if not (-90 <= dest_lat <= 90) or not (-180 <= dest_lon <= 180):
            return jsonify({'error': 'Invalid destination coordinates'}), 400
    except (ValueError, TypeError):
        return jsonify({'error': 'Invalid coordinate format'}), 400
    
    # Validate distance and fare
    try:
        distance_km = Decimal(str(data['distance_km']))
        fare = Decimal(str(data['fare']))
        
        if distance_km <= 0 or distance_km > 1000:
            return jsonify({'error': 'Invalid distance'}), 400
        if fare <= 0 or fare > 100000:
            return jsonify({'error': 'Invalid fare amount'}), 400
    except (ValueError, TypeError):
        return jsonify({'error': 'Invalid distance or fare format'}), 400
    
    # Validate vehicle type
    vehicle_type = data.get('vehicle_type', 'Bajaj')
    if vehicle_type not in ['Bajaj', 'Car']:
        vehicle_type = 'Bajaj'
    
    # Validate payment method
    payment_method = data.get('payment_method', 'Cash')
    if payment_method not in ['Cash', 'Mobile Money', 'Card']:
        payment_method = 'Cash'
    
    # Sanitize optional fields
    pickup_address = data.get('pickup_address', '')[:255] if data.get('pickup_address') else None
    dest_address = data.get('dest_address', '')[:255]
    note = data.get('note', '')[:255] if data.get('note') else None
    
    # Use service layer to create ride
    new_ride = ride_service.create_ride(
        passenger_id=current_user.id,
        pickup_address=pickup_address,
        pickup_lat=pickup_lat,
        pickup_lon=pickup_lon,
        dest_address=dest_address,
        dest_lat=dest_lat,
        dest_lon=dest_lon,
        distance_km=distance_km,
        fare=fare,
        vehicle_type=vehicle_type,
        payment_method=payment_method,
        note=note
    )
    ride_service.commit()

    return jsonify({'message': 'Ride requested successfully', 'ride_id': new_ride.id}), 201


@app.route('/api/assign-ride', methods=['POST'])
@admin_required
def assign_ride():
    data = request.json
    ride = ride_service.get_by_id(data.get('ride_id'))
    driver = driver_service.get_by_id(data.get('driver_id'))
    if not ride or not driver:
        return jsonify({'error': 'Ride or Driver not found'}), 404

    # Use service layer to assign driver
    ride_service.assign_driver(ride, driver.id)
    driver_service.update(driver, status='On Trip', current_lat=ride.pickup_lat, current_lon=ride.pickup_lon)
    ride_service.commit()
    return jsonify({'message': 'Ride assigned successfully'})


@app.route('/api/complete-ride', methods=['POST'])
@admin_required
def complete_ride():
    ride = Ride.query.get(request.json.get('ride_id'))
    if not ride:
        return jsonify({'error': 'Ride not found'}), 404
    ride.status = 'Completed'
    if ride.driver:
        ride.driver.status = 'Available'
    db.session.commit()
    return jsonify({'message': 'Ride marked as completed'})


@app.route('/api/cancel-ride', methods=['POST'])
@login_required # Accessible by both for now, can be split if needed
def cancel_ride():
    ride = Ride.query.get(request.json.get('ride_id'))
    if not ride:
        return jsonify({'error': 'Ride not found'}), 404
    
    # Security check: only allow passenger who owns ride or an admin to cancel
    if session.get('user_type') == 'passenger' and ride.passenger_id != current_user.id:
        return jsonify({'error': 'Unauthorized'}), 403

    is_reassign = ride.status in ['Assigned', 'On Trip'] and session.get('user_type') == 'admin'

    if ride.driver:
        ride.driver.status = 'Available'

    if is_reassign:
        ride.status = 'Requested'
        ride.driver_id = None
        ride.assigned_time = None
        message = "Ride has been reassigned to the pending queue."
    else:
        ride.status = 'Canceled'
        message = 'Ride canceled successfully'

    db.session.commit()
    return jsonify({'message': message})


@app.route('/api/add-driver', methods=['POST'])
@csrf.exempt
@admin_required
def add_driver():
    try:
        # Validate required fields (matching HTML form requirements)
        required_fields = ['name', 'phone_number', 'vehicle_type', 'vehicle_details', 'vehicle_plate_number', 'license_info']
        for field in required_fields:
            if not request.form.get(field, '').strip():
                return jsonify({'error': f'Missing required field: {field}'}), 400
        
        # Validate phone number format
        phone = request.form.get('phone_number', '').strip()
        if not phone or len(phone) < 10:
            return jsonify({'error': 'Invalid phone number format'}), 400
        
        # Check for duplicate phone number
        existing_driver = Driver.query.filter_by(phone_number=phone).first()
        if existing_driver:
            return jsonify({'error': 'Phone number already registered to another driver'}), 409
        
        # Handle file uploads with error checking
        try:
            profile_picture = _handle_file_upload(request.files.get('profile_picture'), 'static/img/default_user.svg')
            license_document = _handle_file_upload(request.files.get('license_document'))
            vehicle_document = _handle_file_upload(request.files.get('vehicle_document'))
        except ValueError as e:
            return jsonify({'error': str(e)}), 400
        
        new_driver = Driver(
            name=request.form.get('name').strip(),
            phone_number=phone,
            vehicle_type=request.form.get('vehicle_type'),
            vehicle_details=request.form.get('vehicle_details').strip(),
            vehicle_plate_number=request.form.get('vehicle_plate_number').strip(),
            license_info=request.form.get('license_info').strip(),
            status='Offline',
            profile_picture=profile_picture,
            license_document=license_document,
            vehicle_document=vehicle_document
        )
        
        db.session.add(new_driver)
        db.session.flush()
        new_driver.driver_uid = f"DRV-{new_driver.id:04d}"
        db.session.commit()
        return jsonify({'message': 'Driver added successfully', 'driver_id': new_driver.id}), 201
        
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Error adding driver: {str(e)}")
        return jsonify({'error': 'Failed to add driver. Please check your input and try again.'}), 500

@app.route('/api/update-driver/<int:driver_id>', methods=['POST'])
@admin_required
def update_driver(driver_id):
    driver = Driver.query.get_or_404(driver_id)
    driver.name = request.form.get('name', driver.name)
    driver.phone_number = request.form.get('phone_number', driver.phone_number)
    driver.vehicle_type = request.form.get('vehicle_type', driver.vehicle_type)
    driver.vehicle_details = request.form.get('vehicle_details', driver.vehicle_details)
    driver.vehicle_plate_number = request.form.get('vehicle_plate_number', driver.vehicle_plate_number)
    driver.license_info = request.form.get('license_info', driver.license_info)
    
    driver.profile_picture = _handle_file_upload(request.files.get('profile_picture'), driver.profile_picture)
    driver.license_document = _handle_file_upload(request.files.get('license_document'), driver.license_document)
    driver.vehicle_document = _handle_file_upload(request.files.get('vehicle_document'), driver.vehicle_document)

    db.session.commit()
    return jsonify({'message': 'Driver updated successfully'})


@app.route('/api/delete-driver', methods=['POST'])
@admin_required
def delete_driver():
    data = request.json
    driver = Driver.query.get(data.get('driver_id'))
    if not driver:
        return jsonify({'error': 'Driver not found'}), 404
    db.session.delete(driver)
    db.session.commit()
    return jsonify({'message': 'Driver deleted successfully'})


@app.route('/api/update-driver-status', methods=['POST'])
@admin_required
def update_driver_status():
    data = request.json
    driver = Driver.query.get(data.get('driver_id'))
    if not driver:
        return jsonify({'error': 'Driver not found'}), 404
    driver.status = data.get('status')
    db.session.commit()
    return jsonify({'message': f'Driver status updated to {driver.status}'})


@app.route('/api/rate-ride', methods=['POST'])
@passenger_required
def rate_ride():
    data = request.json
    # Use service layer to get ride
    ride = ride_service.get_by_id(data.get('ride_id'))
    if not ride:
        return jsonify({'error': 'Ride not found'}), 404
    if ride.passenger_id != current_user.id:
        return jsonify({'error': 'Unauthorized'}), 403
    if ride.status != 'Completed':
        return jsonify({'error': 'Only completed rides can be rated'}), 400
    
    # Use service layer to create or update feedback
    rating = data.get('rating')
    comment = data.get('comment')
    feedback_service.create_or_update_feedback(
        ride_id=ride.id,
        feedback_type='Rating',
        rating=rating,
        comment=comment
    )
    feedback_service.commit()
    return jsonify({'message': 'Thank you for your feedback!'})

@app.route('/api/submit-support-ticket', methods=['POST'])
@passenger_required
def submit_support_ticket():
    data = request.json
    feedback_type = data.get('feedback_type')
    details = data.get('details')
    ride_id = data.get('ride_id')

    if not feedback_type or not details:
        return jsonify({'error': 'Type and details are required.'}), 400
    
    if ride_id:
        # Use service layer to validate ride
        ride = ride_service.get_passenger_rides(current_user.id)
        ride = next((r for r in ride if r.id == ride_id), None)
        if not ride:
            return jsonify({'error': 'Invalid ride ID.'}), 404
        
        # Use service layer to create or update feedback
        # Note: The current DB schema has a UNIQUE constraint on ride_id in Feedback.
        # This means one ride can only have one feedback entry.
        feedback = feedback_service.get_by_ride_id(ride_id)
        if feedback:
            # If feedback already exists, update it by appending the new information
            feedback_service.update(
                feedback,
                feedback_type=f"{feedback.feedback_type}/{feedback_type}",
                details=f"Support Ticket: {details}\n\nOriginal Feedback: {feedback.details or feedback.comment or ''}",
                is_resolved=False
            )
        else:
            # Create new feedback entry using service layer
            feedback_service.create(
                ride_id=ride_id,
                feedback_type=feedback_type,
                details=details,
                is_resolved=False
            )
        feedback_service.commit()
    else:
        # General feedback not linked to a ride - needs a dummy ride_id or schema change
        # This part is tricky with the current schema (ride_id is NOT NULL and FOREIGN KEY).
        # For general feedback without a ride, we cannot store it due to schema constraints.
        return jsonify({'error': 'For general feedback, please select your most recent ride if applicable or contact support directly.'}), 400

    return jsonify({'message': 'Support ticket submitted successfully.'}), 201


@app.route('/api/all-feedback')
@admin_required
def get_all_feedback():
    feedback_items = Feedback.query.order_by(Feedback.is_resolved.asc(), Feedback.submitted_at.desc()).all()
    return jsonify([
        {
            'id': f.id,
            'ride_id': f.ride_id,
            'passenger_name': f.ride.passenger.username,
            'driver_name': f.ride.driver.name if f.ride.driver else 'N/A',
            'rating': f.rating,
            'comment': f.comment,
            'type': f.feedback_type,
            'details': f.details,
            'is_resolved': f.is_resolved,
            'date': to_eat(f.submitted_at).strftime('%Y-%m-%d %H:%M')
        }
        for f in feedback_items
    ])

@app.route('/api/unread-feedback-count')
@admin_required
def get_unread_feedback_count():
    count = Feedback.query.filter_by(is_resolved=False).count()
    return jsonify({'count': count})

@app.route('/api/feedback/resolve/<int:feedback_id>', methods=['POST'])
@admin_required
def resolve_feedback(feedback_id):
    feedback = Feedback.query.get_or_404(feedback_id)
    feedback.is_resolved = True
    db.session.commit()
    return jsonify({'message': 'Feedback marked as resolved.'})


# --- Data Fetching API Routes ---
@app.route('/api/pending-rides')
@admin_required
def get_pending_rides():
    # Use service layer to get pending rides
    rides = ride_service.get_pending_rides()
    return jsonify(rides_schema.dump(rides))

@app.route('/api/active-rides')
@admin_required
def get_active_rides():
    # Use service layer to get active rides
    rides = ride_service.get_active_rides()
    return jsonify([ { 
        'id': r.id, 
        'user_name': r.passenger.username, 
        'driver_name': r.driver.name if r.driver else "N/A", 
        'dest_address': r.dest_address, 
        'status': r.status, 
        'request_time': to_eat(r.request_time).strftime('%Y-%m-%d %H:%M'),
        'pickup_lat': r.pickup_lat,
        'pickup_lon': r.pickup_lon,
        'dest_lat': r.dest_lat,
        'dest_lon': r.dest_lon
    } for r in rides ])


@app.route('/api/drivers')
@admin_required
def get_all_drivers():
    drivers = Driver.query.all()
    drivers_data = []
    for d in drivers:
        avg_rating = db.session.query(func.avg(Feedback.rating)).join(Ride).filter(Ride.driver_id == d.id, Feedback.rating.isnot(None)).scalar() or 0
        d.avg_rating = avg_rating # Add the calculated field to the object
        drivers_data.append(d)
    return jsonify(drivers_schema.dump(drivers_data))


@app.route('/api/driver/<int:driver_id>')
@admin_required
def get_driver(driver_id):
    driver = Driver.query.get_or_404(driver_id)
    return jsonify({ "id": driver.id, "name": driver.name, "phone_number": driver.phone_number, "vehicle_type": driver.vehicle_type, "vehicle_details": driver.vehicle_details, "vehicle_plate_number": driver.vehicle_plate_number, "license_info": driver.license_info, "profile_picture": driver.profile_picture, "license_document": driver.license_document, "vehicle_document": driver.vehicle_document, })


@app.route('/api/available-drivers')
@admin_required
def get_available_drivers():
    vehicle_type = request.args.get('vehicle_type')
    # Use service layer to get available drivers
    drivers = driver_service.get_available_drivers(vehicle_type)
    
    drivers_data = [
        {'id': d.id, 'name': d.name, 'vehicle_type': d.vehicle_type, 'status': d.status}
        for d in drivers
    ]

    return jsonify(drivers_data)


@app.route('/api/all-rides-data')
@admin_required
def get_all_rides_data():
    rides = Ride.query.options(db.joinedload(Ride.feedback)).order_by(Ride.request_time.desc()).all()
    return jsonify([
        {
            'id': r.id,
            'user_name': r.passenger.username,
            'user_phone': r.passenger.phone_number,
            'driver_name': r.driver.name if r.driver else "N/A",
            'fare': r.fare,
            'status': r.status,
            'rating': r.feedback.rating if r.feedback else None,
            'request_time': to_eat(r.request_time).strftime('%Y-%m-%d %H:%M')
        }
        for r in rides
    ])

@app.route('/api/passengers')
@admin_required
def get_passengers():
    passengers = Passenger.query.options(db.selectinload(Passenger.rides)).all()
    return jsonify([
        {
            "id": p.id,
            "username": p.username,
            "phone_number": p.phone_number,
            "profile_picture": p.profile_picture,
            "rides_taken": len(p.rides),
            "join_date": p.join_date.strftime('%Y-%m-%d')
        } for p in passengers
    ])


@app.route('/api/ride-status/<int:ride_id>')
@login_required # Checked inside
def get_ride_status(ride_id):
    ride = Ride.query.get_or_404(ride_id)
    if session.get('user_type') == 'passenger' and ride.passenger_id != current_user.id:
        return jsonify({'error': 'Unauthorized'}), 403
        
    driver_info = None
    if ride.driver:
        driver_info = { 'id': ride.driver.id, 'name': ride.driver.name, 'phone_number': ride.driver.phone_number, 'vehicle_details': ride.driver.vehicle_details }
    ride_details = {'fare': ride.fare, 'dest_address': ride.dest_address} if ride.status == 'Completed' else None
    return jsonify({'status': ride.status, 'driver': driver_info, 'ride_details': ride_details})


@app.route('/api/passenger-details/<int:passenger_id>')
@admin_required
def get_passenger_details(passenger_id):
    passenger = Passenger.query.get_or_404(passenger_id)

    total_spent = db.session.query(func.sum(Ride.fare)).filter(
        Ride.passenger_id == passenger_id,
        Ride.status == 'Completed'
    ).scalar() or 0

    avg_rating_given = db.session.query(func.avg(Feedback.rating)).join(Ride).filter(
        Ride.passenger_id == passenger_id,
        Feedback.rating.isnot(None)
    ).scalar() or 0

    stats = {
        'total_rides': len(passenger.rides),
        'total_spent': total_spent,
        'avg_rating_given': avg_rating_given
    }

    history = Ride.query.filter_by(passenger_id=passenger_id)\
        .options(db.joinedload(Ride.driver), db.joinedload(Ride.feedback))\
        .order_by(Ride.request_time.desc()).limit(20).all()

    return jsonify({
        'profile': {
            'name': passenger.username,
            'passenger_uid': passenger.passenger_uid,
            'phone_number': passenger.phone_number,
            'avatar': passenger.profile_picture,
            'join_date': passenger.join_date.strftime('%b %d, %Y')
        },
        'stats': {k: round(v, 2) if isinstance(v, float) else v for k, v in stats.items()},
        'history': [{
            'id': r.id,
            'status': r.status,
            'fare': r.fare,
            'date': to_eat(r.request_time).strftime('%Y-%m-%d %H:%M'),
            'driver_name': r.driver.name if r.driver else 'N/A',
            'pickup_address': r.pickup_address,
            'dest_address': r.dest_address,
            'rating_given': r.feedback.rating if r.feedback and r.feedback.rating is not None else 'N/A'
        } for r in history]
    })


@app.route('/api/ride-details/<int:ride_id>')
@admin_required
def get_ride_details(ride_id):
    ride = Ride.query.options(
        db.joinedload(Ride.passenger),
        db.joinedload(Ride.driver),
        db.joinedload(Ride.feedback)
    ).get_or_404(ride_id)

    return jsonify({
        'trip_info': {
            'id': ride.id,
            'status': ride.status,
            'fare': ride.fare,
            'distance': ride.distance_km,
            'payment_method': ride.payment_method,
            'vehicle_type': ride.vehicle_type,
            'pickup_address': ride.pickup_address,
            'dest_address': ride.dest_address,
            'pickup_coords': {'lat': ride.pickup_lat, 'lon': ride.pickup_lon},
            'dest_coords': {'lat': ride.dest_lat, 'lon': ride.dest_lon},
        },
        'passenger': {
            'name': ride.passenger.username,
            'avatar': ride.passenger.profile_picture,
            'phone': ride.passenger.phone_number
        },
        'driver': {
            'name': ride.driver.name if ride.driver else 'N/A',
            'avatar': ride.driver.profile_picture if ride.driver else 'static/img/default_user.svg',
            'phone': ride.driver.phone_number if ride.driver else 'N/A',
            'vehicle': ride.driver.vehicle_details if ride.driver else 'N/A'
        },
        'timestamps': {
            'requested': to_eat(ride.request_time).strftime('%b %d, %Y at %I:%M %p'),
            'assigned': to_eat(ride.assigned_time).strftime('%I:%M %p') if ride.assigned_time else 'N/A',
        },
        'feedback': {
            'rating': ride.feedback.rating if ride.feedback else None,
            'comment': ride.feedback.comment if ride.feedback else None
        }
    })


@app.route('/api/driver-details/<int:driver_id>')
@admin_required
def get_driver_details(driver_id):
    driver = Driver.query.get_or_404(driver_id)
    
    now = datetime.now(timezone.utc)
    week_start = now - timedelta(days=now.weekday())

    total_earnings_all_time = db.session.query(func.sum(Ride.fare)).filter(Ride.driver_id == driver_id, Ride.status == 'Completed').scalar() or 0
    stats = {
        'completed_rides': Ride.query.filter_by(driver_id=driver_id, status='Completed').count(),
        'total_earnings_all_time': total_earnings_all_time,
        'total_earnings_weekly': db.session.query(func.sum(Ride.fare)).filter(Ride.driver_id == driver_id, Ride.status == 'Completed', Ride.request_time >= week_start).scalar() or 0,
        'avg_rating': db.session.query(func.avg(Feedback.rating)).join(Ride).filter(Ride.driver_id == driver.id, Feedback.rating.isnot(None)).scalar() or 0
    }
    history = Ride.query.filter_by(driver_id=driver_id).order_by(Ride.request_time.desc()).limit(10).all()
    return jsonify({
        'profile': { 'name': driver.name, 'driver_uid': driver.driver_uid, 'status': driver.status, 'avatar': driver.profile_picture, 'phone_number': driver.phone_number, 'vehicle_type': driver.vehicle_type, 'vehicle_details': driver.vehicle_details, 'plate_number': driver.vehicle_plate_number, 'license': driver.license_info, 'license_document': driver.license_document, 'vehicle_document': driver.vehicle_document },
        'stats': {k: round(v, 2) if isinstance(v, float) else v for k, v in stats.items()},
        'history': [{'id': r.id, 'status': r.status, 'fare': r.fare, 'date': to_eat(r.request_time).strftime('%Y-%m-%d')} for r in history]
    })

@app.route('/api/passenger/ride-history')
@passenger_required
def api_passenger_history():
    rides = Ride.query.filter_by(passenger_id=current_user.id)\
        .options(db.joinedload(Ride.feedback), db.joinedload(Ride.driver))\
        .order_by(Ride.request_time.desc()).all()
    
    return jsonify([
        {
            'id': r.id,
            'driver_name': r.driver.name if r.driver else "N/A",
            'dest_address': r.dest_address,
            'fare': r.fare,
            'status': r.status,
            'request_time': to_eat(r.request_time).strftime('%b %d, %Y at %I:%M %p'),
            'rating': r.feedback.rating if r.feedback else None,
            'comment': r.feedback.comment if r.feedback else None
        } for r in rides
    ])


@app.route('/api/fare-estimate', methods=['POST'])
@csrf.exempt
@limiter.exempt  # Allow frequent fare estimates
def fare_estimate():
    try:
        # Debug logging
        app.logger.info(f"Fare estimate request - Content-Type: {request.content_type}")
        app.logger.info(f"Fare estimate request - Raw data: {request.get_data()}")
        
        data = request.json
        app.logger.info(f"Parsed JSON data: {data}")
        
        if not data:
            app.logger.error("No JSON data provided")
            return jsonify({'error': 'No data provided'}), 400
        
        # Validate required fields
        required_fields = ['pickup_lat', 'pickup_lon', 'dest_lat', 'dest_lon']
        for field in required_fields:
            if field not in data:
                return jsonify({'error': f'Missing required field: {field}'}), 400
    
        # Validate coordinate values
        try:
            pickup_lat = float(data['pickup_lat'])
            pickup_lon = float(data['pickup_lon'])
            dest_lat = float(data['dest_lat'])
            dest_lon = float(data['dest_lon'])
            
            # Basic coordinate validation
            if not (-90 <= pickup_lat <= 90) or not (-180 <= pickup_lon <= 180):
                return jsonify({'error': 'Invalid pickup coordinates'}), 400
            if not (-90 <= dest_lat <= 90) or not (-180 <= dest_lon <= 180):
                return jsonify({'error': 'Invalid destination coordinates'}), 400
        except (ValueError, TypeError):
            return jsonify({'error': 'Invalid coordinate format'}), 400
        
        # Get pricing settings
        base_fare = Decimal(get_setting('base_fare', '25'))
        per_km_rates = {
            "Bajaj": Decimal(get_setting('per_km_bajaj', '8')),
            "Car": Decimal(get_setting('per_km_car', '12'))
        }
        vehicle_type = data.get('vehicle_type', 'Bajaj')
        per_km_rate = per_km_rates.get(vehicle_type, per_km_rates['Bajaj'])
        
        # Call OSRM routing service
        osrm_url = (f"http://router.project-osrm.org/route/v1/driving/"
                    f"{pickup_lon},{pickup_lat};{dest_lon},{dest_lat}?overview=false")
        
        try:
            response = requests.get(osrm_url, timeout=15)
            response.raise_for_status()
            
            route_data = response.json()
            if 'routes' not in route_data or not route_data['routes']:
                return jsonify({'error': 'No route found between the locations'}), 400
            
            distance_meters = route_data['routes'][0]['distance']
            distance_km = Decimal(str(distance_meters / 1000.0))
            
        except requests.exceptions.Timeout:
            return jsonify({'error': 'Route calculation timed out. Please try again.'}), 503
        except requests.exceptions.ConnectionError:
            return jsonify({'error': 'Unable to connect to routing service. Please try again later.'}), 503
        except requests.exceptions.RequestException as e:
            app.logger.error(f"OSRM request failed: {str(e)}")
            return jsonify({'error': 'Route service temporarily unavailable'}), 503
        except (KeyError, IndexError, TypeError) as e:
            app.logger.error(f"OSRM response parsing failed: {str(e)}")
            return jsonify({'error': 'Invalid route response'}), 500
        
        # Calculate fare
        fare = (base_fare + (distance_km * per_km_rate)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        
        return jsonify({
            'distance_km': float(distance_km.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)),
            'estimated_fare': float(fare)
        })
        
    except Exception as e:
        app.logger.error(f"Fare estimation error: {str(e)}")
        return jsonify({'error': 'Unable to calculate fare. Please try again.'}), 500


@app.route('/api/dashboard-stats')
@admin_required
def get_dashboard_stats():
    total_revenue = db.session.query(func.sum(Ride.fare)).filter(Ride.status == 'Completed').scalar() or 0
    total_rides = Ride.query.count()
    drivers_online = Driver.query.filter(Driver.status == 'Available').count()
    pending_requests = Ride.query.filter(Ride.status == 'Requested').count()
    return jsonify({ 'total_revenue': round(total_revenue, 2), 'total_rides': total_rides, 'drivers_online': drivers_online, 'pending_requests': pending_requests, 'active_rides': Ride.query.filter(Ride.status.in_(['Assigned', 'On Trip'])).count() })

# --- Analytics and Reporting ---
def _get_previous_period(start_date, end_date):
    if not start_date or not end_date: return None, None
    delta = end_date - start_date
    prev_end_date = start_date - timedelta(microseconds=1)
    prev_start_date = prev_end_date - delta
    return prev_start_date, prev_end_date

def _calculate_kpis_for_period(start, end):
    query = Ride.query
    if start and end: query = query.filter(Ride.request_time.between(start, end))
    completed_sq = query.filter(Ride.status == 'Completed').subquery()
    revenue = db.session.query(func.sum(completed_sq.c.fare)).scalar() or 0
    completed_rides = db.session.query(func.count(completed_sq.c.id)).scalar()
    return revenue, completed_rides

def _calculate_trend(current, previous):
    if previous == 0: return 100 if current > 0 else 0
    return round(((current - previous) / previous) * 100)
    
def _get_date_range_from_request():
    period = request.args.get('period')
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')

    if period == 'all' or not period:
        return None, None

    now = datetime.now(timezone.utc)

    if period == 'today':
        start_date = now.replace(hour=0, minute=0, second=0, microsecond=0)
        end_date = now
    elif period == 'week':
        start_date = (now - timedelta(days=now.weekday())).replace(hour=0, minute=0, second=0, microsecond=0)
        end_date = now
    elif period == 'month':
        start_date = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        end_date = now
    elif start_date_str and end_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').replace(tzinfo=timezone.utc)
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').replace(hour=23, minute=59, second=59, tzinfo=timezone.utc)
        except (ValueError, TypeError):
            return None, None
    else:
        return None, None

    return start_date, end_date

@app.route('/api/analytics-data')
@admin_required
def get_analytics_data():
    start_date, end_date = _get_date_range_from_request()
    now = datetime.now(timezone.utc)
    base_query = Ride.query
    if start_date and end_date:
        base_query = base_query.filter(Ride.request_time.between(start_date, end_date))

    completed_rides_sq = base_query.filter(Ride.status == 'Completed').subquery()
    completed_rides_in_period = db.session.query(func.count(completed_rides_sq.c.id)).scalar() or 0
    revenue_in_period = db.session.query(func.sum(completed_rides_sq.c.fare)).scalar() or 0
    prev_start_date, prev_end_date = _get_previous_period(start_date, end_date)
    prev_revenue, prev_completed_rides = (0, 0) if not (prev_start_date and prev_end_date) else _calculate_kpis_for_period(prev_start_date, prev_end_date)
    
    now_week_start = (now - timedelta(days=now.weekday())).replace(hour=0, minute=0, second=0, microsecond=0)
    
    top_drivers_query = db.session.query(
        Driver.name,
        Driver.profile_picture,
        func.count(Ride.id).label('completed_rides'),
        func.avg(Feedback.rating).label('avg_rating')
    ).join(Ride, Driver.id == Ride.driver_id).outerjoin(Feedback, Ride.id == Feedback.ride_id).filter(
        Ride.status == 'Completed',
        Ride.request_time >= now_week_start
    ).group_by(Driver.id).order_by(
        func.count(Ride.id).desc()
    ).limit(5).all()

    return jsonify({
        'kpis': { 'rides_completed': completed_rides_in_period, 'rides_canceled': base_query.filter(Ride.status == 'Canceled').count(), 'total_revenue': round(revenue_in_period, 2), 'avg_fare': round(db.session.query(func.avg(completed_rides_sq.c.fare)).scalar() or 0, 2), 'active_rides_now': Ride.query.filter(Ride.status.in_(['Assigned', 'On Trip'])).count(), 'trends': { 'revenue': _calculate_trend(revenue_in_period, prev_revenue), 'rides': _calculate_trend(completed_rides_in_period, prev_completed_rides) } },
        'charts': { 
            'revenue_over_time': { 
                'labels': [i[0].strftime('%Y-%m-%d') if i[0] else '' for i in db.session.query(func.date(completed_rides_sq.c.request_time).label('d'), func.sum(completed_rides_sq.c.fare)).group_by('d').order_by('d').all()], 
                'data': [float(i[1] or 0) for i in db.session.query(func.date(completed_rides_sq.c.request_time).label('d'), func.sum(completed_rides_sq.c.fare)).group_by('d').order_by('d').all()] 
            }, 
            'vehicle_distribution': dict(base_query.with_entities(Ride.vehicle_type, func.count(Ride.id)).group_by(Ride.vehicle_type).all()), 
            'payment_method_distribution': dict(base_query.with_entities(Ride.payment_method, func.count(Ride.id)).group_by(Ride.payment_method).all())
        },
        'performance': { 'top_drivers': [{'name': d.name, 'avatar': d.profile_picture, 'completed_rides': d.completed_rides, 'avg_rating': round(d.avg_rating or 0, 2)} for d in top_drivers_query] }
    })

@app.route('/api/export-report')
@admin_required
def export_report():
    file_format = request.args.get('format', 'pdf')
    start_date, end_date = _get_date_range_from_request()
    
    ride_query = Ride.query.options(db.joinedload(Ride.passenger), db.joinedload(Ride.driver)).order_by(Ride.request_time.desc())
    if start_date and end_date:
        ride_query = ride_query.filter(Ride.request_time.between(start_date, end_date))
    
    all_rides_in_period = ride_query.all()
    all_drivers = Driver.query.all()
    
    with app.test_request_context(f'/api/analytics-data?{request.query_string.decode("utf-8")}'):
        kpis_response = get_analytics_data()
        kpis = kpis_response.get_json().get('kpis', {}) if kpis_response.get_json() else {}

    report_title = f"Analytics Report ({to_eat(start_date).strftime('%Y-%m-%d')} to {to_eat(end_date).strftime('%Y-%m-%d')})" if start_date and end_date else "Analytics Report (All Time)"

    if file_format == 'pdf':
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
        styles = getSampleStyleSheet()
        elements = [ Paragraph("Ride App - Dispatcher Analytics", styles['h1']), Paragraph(report_title, styles['h2']), Spacer(1, 24) ]
        kpi_data = [ ["Metric", "Value", "Trend"], ["Rides Completed", f"{kpis['rides_completed']}", f"{kpis['trends']['rides']}%"], ["Total Revenue", f"{kpis['total_revenue']} ETB", f"{kpis['trends']['revenue']}%"], ["Rides Canceled", kpis['rides_canceled'], ""], ["Average Fare", f"{kpis['avg_fare']} ETB", ""], ]
        kpi_table = Table(kpi_data, colWidths=[200, 150, 100]); kpi_table.setStyle(TableStyle([ ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#4A5568')), ('TEXTCOLOR',(0,0),(-1,0),colors.whitesmoke), ('ALIGN', (0,0), (-1,-1), 'CENTER'), ('VALIGN', (0,0), (-1,-1), 'MIDDLE'), ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'), ('GRID', (0,0), (-1,-1), 1, colors.black) ])); elements.extend([Paragraph("Key Metrics Summary", styles['h3']), kpi_table, Spacer(1, 24)])
        ride_data_for_table = [["ID", "Date", "Passenger", "Driver", "Fare", "Status"]]
        for ride in all_rides_in_period: ride_data_for_table.append([ ride.id, to_eat(ride.request_time).strftime('%Y-%m-%d %H:%M'), ride.passenger.username, ride.driver.name if ride.driver else 'N/A', f"{ride.fare} ETB", ride.status ])
        ride_table = Table(ride_data_for_table, colWidths=[40, 120, 120, 120, 80, 80]); ride_table.setStyle(TableStyle([ ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#2D3748')), ('TEXTCOLOR',(0,0),(-1,0),colors.whitesmoke), ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'), ('GRID', (0,0), (-1,-1), 1, colors.black), ('ALIGN', (0,0), (-1,-1), 'CENTER'), ])); elements.extend([Paragraph("All Rides in Period", styles['h3']), ride_table]); doc.build(elements); buffer.seek(0)
        return buffer.getvalue(), 200, { 'Content-Type': 'application/pdf', 'Content-Disposition': 'attachment; filename="analytics_report.pdf"' }

    elif file_format == 'excel':
        wb = openpyxl.Workbook()
        ws_summary = wb.active
        ws_summary.title = "Summary Report"
        ws_summary.append([report_title])
        ws_summary['A1'].font = Font(bold=True, size=16)
        ws_summary.append([])
        ws_summary.append(["Key Metrics"])
        ws_summary['A3'].font = Font(bold=True, size=14)
        ws_summary.append(["Metric", "Value"])
        for cell in ws_summary[4]:
            cell.font = Font(bold=True)
        ws_summary.append(["Rides Completed", kpis.get('rides_completed', 'N/A')])
        ws_summary.append(["Rides Canceled", kpis.get('rides_canceled', 'N/A')])
        ws_summary.append(["Total Revenue (ETB)", kpis.get('total_revenue', 'N/A')])
        ws_summary.append(["Average Fare (ETB)", kpis.get('avg_fare', 'N/A')])
        ws_summary['D4'] = "Rides Trend"; ws_summary['D4'].font = Font(bold=True)
        ws_summary['E4'] = f"{kpis.get('trends', {}).get('rides', 'N/A')}%"
        ws_summary['D5'] = "Revenue Trend"; ws_summary['D5'].font = Font(bold=True)
        ws_summary['E5'] = f"{kpis.get('trends', {}).get('revenue', 'N/A')}%"
        
        ws_drivers = wb.create_sheet("Driver Performance")
        ws_drivers.append(["Driver ID", "Name", "Rides in Period", "Revenue in Period (ETB)", "Average Rating"])
        for cell in ws_drivers[1]:
            cell.font = Font(bold=True)
        for driver in all_drivers:
            rides_in_period_query = Ride.query.filter(Ride.driver_id == driver.id, Ride.status == 'Completed')
            if start_date and end_date: 
                rides_in_period_query = rides_in_period_query.filter(Ride.request_time.between(start_date, end_date))
            
            rides_in_period_count = rides_in_period_query.count()
            total_revenue = rides_in_period_query.with_entities(func.sum(Ride.fare)).scalar() or 0
            avg_rating = db.session.query(func.avg(Feedback.rating)).join(Ride).filter(Ride.driver_id == driver.id, Feedback.rating.isnot(None)).scalar() or 0
            ws_drivers.append([ driver.driver_uid, driver.name, rides_in_period_count, round(total_revenue, 2), round(avg_rating, 2) if avg_rating else 0 ])

        ws_raw = wb.create_sheet("Raw Ride Data")
        ws_raw.append(["Ride ID", "Request Time (EAT)", "Status", "Passenger Name", "Passenger Phone", "Driver Name", "Fare", "Vehicle", "Payment", "Rating"])
        for cell in ws_raw[1]:
            cell.font = Font(bold=True)
        for ride in all_rides_in_period: ws_raw.append([ ride.id, to_eat(ride.request_time).strftime('%Y-%m-%d %H:%M'), ride.status, ride.passenger.username, ride.passenger.phone_number, ride.driver.name if ride.driver else 'N/A', ride.fare, ride.vehicle_type, ride.payment_method, ride.feedback.rating if ride.feedback else None ])
        
        for ws in wb.worksheets:
            for col in ws.columns:
                max_length = 0
                column_letter = get_column_letter(col[0].column)
                for cell in col:
                    try: 
                        if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
                    except: pass
                ws.column_dimensions[column_letter].width = (max_length + 2) if max_length < 50 else 50
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue(), 200, { 'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 'Content-Disposition': 'attachment; filename="analytics_report.xlsx"' }
    return jsonify({"error": "Invalid format"}), 400


@app.route('/api/settings', methods=['GET', 'POST'])
@admin_required
def handle_settings():
    if request.method == 'POST':
        for key, value in request.json.items():
            setting = Setting.query.filter_by(key=key).first();
            if setting: setting.value = str(value)
            else: db.session.add(Setting(key=key, value=str(value)))
        db.session.commit(); return jsonify({'message': 'Settings saved successfully!'})
    else: return jsonify({s.key: s.value for s in Setting.query.all()})

# --- Admin Management API ---
@app.route('/api/admins', methods=['GET'])
@admin_required
def get_admins():
    admins = Admin.query.all()
    return jsonify([{'id': admin.id, 'username': admin.username} for admin in admins])

@app.route('/api/admins/add', methods=['POST'])
@admin_required
def add_admin():
    data = request.json
    username = data.get('username')
    password = data.get('password')

    if not username or not password:
        return jsonify({'error': 'Username and password are required'}), 400
    
    if Admin.query.filter_by(username=username).first():
        return jsonify({'error': 'Username already exists'}), 409

    new_admin = Admin(username=username)
    new_admin.set_password(password)
    db.session.add(new_admin)
    db.session.commit()
    return jsonify({'message': 'Admin added successfully', 'admin': {'id': new_admin.id, 'username': new_admin.username}}), 201

@app.route('/api/admins/delete', methods=['POST'])
@admin_required
def delete_admin():
    data = request.json
    admin_id = data.get('admin_id')
    
    if admin_id == current_user.id:
        return jsonify({'error': 'You cannot delete your own account'}), 403

    if Admin.query.count() <= 1:
        return jsonify({'error': 'Cannot delete the last admin account'}), 403

    admin_to_delete = Admin.query.get(admin_id)
    if not admin_to_delete:
        return jsonify({'error': 'Admin not found'}), 404
        
    db.session.delete(admin_to_delete)
    db.session.commit()
    return jsonify({'message': 'Admin deleted successfully'})

@app.route('/api/admins/update-profile', methods=['POST'])
@admin_required
def update_profile():
    new_username = request.form.get('username')
    current_password = request.form.get('current_password')
    new_password = request.form.get('new_password')
    profile_picture = request.files.get('profile_picture')

    if not current_user.check_password(current_password):
        return jsonify({'error': 'Your current password is not correct'}), 403
    
    if new_username and new_username != current_user.username:
        if Admin.query.filter_by(username=new_username).first():
            return jsonify({'error': 'New username is already taken'}), 409
        current_user.username = new_username
    
    if new_password:
        current_user.set_password(new_password)
    
    if profile_picture:
        current_user.profile_picture = _handle_file_upload(profile_picture, current_user.profile_picture)

    db.session.commit()


# --- Main Execution ---
if __name__ == '__main__':
    with app.app_context():
        os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
        
        if not Admin.query.first():
            print("WARNING: No admin user found. Please create one using the create_admin.py script.")

        if Driver.query.filter(Driver.driver_uid == None).first():
            for driver in Driver.query.filter(Driver.driver_uid == None).all():
                driver.driver_uid = f"DRV-{driver.id:04d}"
            db.session.commit()
            
        if Passenger.query.filter(Passenger.passenger_uid == None).first():
            for p in Passenger.query.filter(Passenger.passenger_uid == None).all():
                p.passenger_uid = f"PAX-{p.id:05d}"
            db.session.commit()
            
        if not Setting.query.first():
            db.session.add(Setting(key='base_fare', value='25'))
            db.session.add(Setting(key='per_km_bajaj', value='8'))
            db.session.add(Setting(key='per_km_car', value='12'))
            db.session.commit()
            
    app.run(debug=True)
