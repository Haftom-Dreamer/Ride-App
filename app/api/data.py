"""
Data retrieval and analytics API endpoints
"""

import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from datetime import datetime, timezone, timedelta
from decimal import Decimal

from flask import request, jsonify, current_app
from sqlalchemy import func, case
from app.models import db, Driver, Ride, Passenger, Feedback, Setting, Admin, DriverEarnings, Commission
from app.api import api, admin_required, passenger_required, get_setting
from app.utils import to_eat
from flask_login import current_user

# --- Dashboard Stats ---
@api.route('/commission-settings', methods=['GET', 'POST'])
@admin_required
def commission_settings():
    """Get or update commission settings"""
    if request.method == 'GET':
        try:
            from app.models import Setting
            
            # Get current commission settings
            bajaj_rate = Setting.query.filter_by(key='commission_rate_bajaj').first()
            car_rate = Setting.query.filter_by(key='commission_rate_car').first()
            
            return jsonify({
                'success': True,
                'settings': {
                    'bajaj_rate': float(bajaj_rate.value) if bajaj_rate else 15.0,
                    'car_rate': float(car_rate.value) if car_rate else 20.0
                }
            })
        except Exception as e:
            return jsonify({'error': str(e)}), 500
    
    elif request.method == 'POST':
        try:
            from app.models import Setting
            
            data = request.json
            bajaj_rate = data.get('bajaj_rate', 15.0)
            car_rate = data.get('car_rate', 20.0)
            
            # Validate rates
            if not (0 <= bajaj_rate <= 100) or not (0 <= car_rate <= 100):
                return jsonify({'error': 'Commission rates must be between 0 and 100'}), 400
            
            # Update or create settings
            bajaj_setting = Setting.query.filter_by(key='commission_rate_bajaj').first()
            if bajaj_setting:
                bajaj_setting.value = str(bajaj_rate)
            else:
                bajaj_setting = Setting(key='commission_rate_bajaj', value=str(bajaj_rate))
                db.session.add(bajaj_setting)
            
            car_setting = Setting.query.filter_by(key='commission_rate_car').first()
            if car_setting:
                car_setting.value = str(car_rate)
            else:
                car_setting = Setting(key='commission_rate_car', value=str(car_rate))
                db.session.add(car_setting)
            
            db.session.commit()
            
            return jsonify({
                'success': True,
                'message': 'Commission settings updated successfully',
                'settings': {
                    'bajaj_rate': bajaj_rate,
                    'car_rate': car_rate
                }
            })
            
        except Exception as e:
            db.session.rollback()
            return jsonify({'error': str(e)}), 500

@api.route('/dashboard-stats')
@admin_required
def get_dashboard_stats():
    """Get dashboard statistics"""
    # Calculate total revenue from completed rides
    total_revenue = db.session.query(func.sum(Ride.fare)).filter(Ride.status == 'Completed').scalar() or 0
    
    # Count all rides
    total_rides = Ride.query.count()
    
    # Count drivers by status
    drivers_online = Driver.query.filter(Driver.status == 'Available').count()
    total_drivers = Driver.query.count()
    
    # Count passengers
    total_passengers = Passenger.query.count()
    
    # Count rides by status
    pending_requests = Ride.query.filter(Ride.status == 'Requested').count()
    active_rides = Ride.query.filter(Ride.status.in_(['Assigned', 'On Trip'])).count()
    completed_rides = Ride.query.filter(Ride.status == 'Completed').count()
    
    # Calculate today's revenue
    today = datetime.now(timezone.utc).date()
    today_start = datetime.combine(today, datetime.min.time()).replace(tzinfo=timezone.utc)
    today_end = datetime.combine(today, datetime.max.time()).replace(tzinfo=timezone.utc)
    
    today_revenue = db.session.query(func.sum(Ride.fare)).filter(
        Ride.status == 'Completed',
        Ride.request_time >= today_start,
        Ride.request_time <= today_end
    ).scalar() or 0
    
    # Count support tickets
    from app.models import SupportTicket
    open_tickets = SupportTicket.query.filter(SupportTicket.status == 'Open').count()
    total_tickets = SupportTicket.query.count()
    
    return jsonify({
        'total_revenue': round(float(total_revenue), 2),
        'total_rides': total_rides,
        'drivers_online': drivers_online,
        'total_drivers': total_drivers,
        'total_passengers': total_passengers,
        'pending_requests': pending_requests,
        'active_rides': active_rides,
        'completed_rides': completed_rides,
        'today_revenue': round(float(today_revenue), 2),
        'open_tickets': open_tickets,
        'total_tickets': total_tickets
    })

# --- Ride Data ---
@api.route('/pending-rides')
@admin_required
def get_pending_rides():
    """Get pending ride requests"""
    rides = Ride.query.filter_by(status='Requested').options(
        db.joinedload(Ride.passenger)
    ).order_by(Ride.request_time.desc()).all()
    
    rides_data = []
    for ride in rides:
        rides_data.append({
            'id': ride.id,
            'user_name': ride.passenger.username,
            'user_phone': ride.passenger.phone_number,
            'pickup_address': ride.pickup_address,
            'pickup_lat': ride.pickup_lat,
            'pickup_lon': ride.pickup_lon,
            'dest_address': ride.dest_address,
            'dest_lat': ride.dest_lat,
            'dest_lon': ride.dest_lon,
            'fare': float(ride.fare),
            'vehicle_type': ride.vehicle_type,
            'note': ride.note,
            'request_time': to_eat(ride.request_time).strftime('%I:%M %p')
        })
    
    return jsonify(rides_data)

@api.route('/active-rides')
@admin_required
def get_active_rides():
    """Get active rides (assigned or in progress)"""
    rides = Ride.query.filter(
        Ride.status.in_(['Assigned', 'On Trip'])
    ).options(
        db.joinedload(Ride.passenger),
        db.joinedload(Ride.driver)
    ).order_by(Ride.request_time.desc()).all()
    
    rides_data = []
    for ride in rides:
        rides_data.append({
            'id': ride.id,
            'user_name': ride.passenger.username,
            'driver_name': ride.driver.name if ride.driver else "N/A",
            'dest_address': ride.dest_address,
            'status': ride.status,
            'request_time': to_eat(ride.request_time).strftime('%Y-%m-%d %H:%M'),
            'pickup_lat': ride.pickup_lat,
            'pickup_lon': ride.pickup_lon,
            'dest_lat': ride.dest_lat,
            'dest_lon': ride.dest_lon
        })
    
    return jsonify(rides_data)

@api.route('/all-rides-data')
@admin_required
def get_all_rides_data():
    """Get all ride data for management"""
    rides = Ride.query.options(
        db.joinedload(Ride.passenger),
        db.joinedload(Ride.driver),
        db.joinedload(Ride.feedback)
    ).order_by(Ride.request_time.desc()).all()
    
    rides_data = []
    for ride in rides:
        rides_data.append({
            'id': ride.id,
            'user_name': ride.passenger.username,
            'user_phone': ride.passenger.phone_number,
            'driver_name': ride.driver.name if ride.driver else "N/A",
            'fare': float(ride.fare),
            'status': ride.status,
            'rating': ride.feedback.rating if ride.feedback else None,
            'request_time': to_eat(ride.request_time).strftime('%Y-%m-%d %H:%M')
        })
    
    return jsonify(rides_data)

@api.route('/ride-details/<int:ride_id>')
@admin_required
def get_ride_details(ride_id):
    """Get detailed information about a specific ride"""
    ride = Ride.query.options(
        db.joinedload(Ride.passenger),
        db.joinedload(Ride.driver),
        db.joinedload(Ride.feedback)
    ).get_or_404(ride_id)

    return jsonify({
        'trip_info': {
            'id': ride.id,
            'status': ride.status,
            'fare': float(ride.fare),
            'distance': float(ride.distance_km),
            'payment_method': ride.payment_method,
            'vehicle_type': ride.vehicle_type,
            'pickup_address': ride.pickup_address,
            'dest_address': ride.dest_address,
            'pickup_coords': {'lat': ride.pickup_lat, 'lon': ride.pickup_lon},
            'dest_coords': {'lat': ride.dest_lat, 'lon': ride.dest_lon},
        },
        'passenger': {
            'name': ride.passenger.username,
            'avatar': ride.passenger.profile_picture,
            'phone': ride.passenger.phone_number
        },
        'driver': {
            'name': ride.driver.name if ride.driver else 'N/A',
            'avatar': ride.driver.profile_picture if ride.driver else 'static/img/default_user.svg',
            'phone': ride.driver.phone_number if ride.driver else 'N/A',
            'vehicle': ride.driver.vehicle_details if ride.driver else 'N/A'
        },
        'timestamps': {
            'requested': to_eat(ride.request_time).strftime('%b %d, %Y at %I:%M %p'),
            'assigned': to_eat(ride.assigned_time).strftime('%I:%M %p') if ride.assigned_time else 'N/A',
        },
        'feedback': {
            'rating': ride.feedback.rating if ride.feedback else None,
            'comment': ride.feedback.comment if ride.feedback else None
        }
    })

# --- Passenger Data ---
@api.route('/passengers')
@admin_required
def get_passengers():
    """Get all passengers"""
    passengers = Passenger.query.options(db.selectinload(Passenger.rides)).all()
    
    passengers_data = []
    for passenger in passengers:
        passengers_data.append({
            "id": passenger.id,
            "passenger_uid": passenger.passenger_uid,
            "username": passenger.username,
            "phone_number": passenger.phone_number,
            "profile_picture": passenger.profile_picture,
            "rides_taken": len(passenger.rides),
            "join_date": passenger.join_date.strftime('%Y-%m-%d') if passenger.join_date else None,
            "is_blocked": passenger.is_blocked,
            "blocked_reason": passenger.blocked_reason if passenger.is_blocked else None
        })
    
    return jsonify(passengers_data)

@api.route('/passenger-details/<int:passenger_id>')
@admin_required
def get_passenger_details(passenger_id):
    """Get detailed passenger information"""
    passenger = Passenger.query.get_or_404(passenger_id)

    # Calculate statistics
    total_spent = db.session.query(func.sum(Ride.fare)).filter(
        Ride.passenger_id == passenger_id,
        Ride.status == 'Completed'
    ).scalar() or 0

    avg_rating_given = db.session.query(func.avg(Feedback.rating)).join(Ride).filter(
        Ride.passenger_id == passenger_id,
        Feedback.rating.isnot(None)
    ).scalar() or 0

    stats = {
        'total_rides': len(passenger.rides),
        'total_spent': float(total_spent),
        'avg_rating_given': round(float(avg_rating_given), 2) if avg_rating_given else 0
    }

    # Get ride history
    history = Ride.query.filter_by(passenger_id=passenger_id)\
        .options(db.joinedload(Ride.driver), db.joinedload(Ride.feedback))\
        .order_by(Ride.request_time.desc()).limit(20).all()

    return jsonify({
        'profile': {
            'id': passenger.id,
            'name': passenger.username,
            'passenger_uid': passenger.passenger_uid,
            'phone_number': passenger.phone_number,
            'avatar': passenger.profile_picture,
            'join_date': passenger.join_date.strftime('%b %d, %Y') if passenger.join_date else None,
            'is_blocked': passenger.is_blocked,
            'blocked_reason': passenger.blocked_reason if passenger.is_blocked else None,
            'blocked_at': to_eat(passenger.blocked_at).strftime('%b %d, %Y') if passenger.blocked_at else None
        },
        'stats': stats,
        'history': [{
            'id': ride.id,
            'status': ride.status,
            'fare': float(ride.fare),
            'date': to_eat(ride.request_time).strftime('%Y-%m-%d %H:%M'),
            'driver_name': ride.driver.name if ride.driver else 'N/A',
            'pickup_address': ride.pickup_address,
            'dest_address': ride.dest_address,
            'rating_given': ride.feedback.rating if ride.feedback and ride.feedback.rating is not None else 'N/A'
        } for ride in history]
    })

@api.route('/passenger/ride-history')
@passenger_required
def api_passenger_history():
    """Get ride history for logged-in passenger"""
    from flask_login import current_user
    
    rides = Ride.query.filter_by(passenger_id=current_user.id)\
        .options(db.joinedload(Ride.feedback), db.joinedload(Ride.driver))\
        .order_by(Ride.request_time.desc()).all()
    
    rides_data = []
    for ride in rides:
        rides_data.append({
            'id': ride.id,
            'driver_name': ride.driver.name if ride.driver else "N/A",
            'dest_address': ride.dest_address,
            'fare': float(ride.fare),
            'status': ride.status,
            'request_time': to_eat(ride.request_time).strftime('%b %d, %Y at %I:%M %p'),
            'rating': ride.feedback.rating if ride.feedback else None,
            'comment': ride.feedback.comment if ride.feedback else None
        })
    
    return jsonify(rides_data)

# --- Feedback and Rating Endpoints ---
@api.route('/rate-ride', methods=['POST'])
@passenger_required
def rate_ride():
    """Submit or update ride rating"""
    try:
        data = request.get_json()
        ride_id = data.get('ride_id')
        rating = data.get('rating')
        comment = data.get('comment', '')
        
        if not ride_id or not rating:
            return jsonify({'error': 'Missing ride_id or rating'}), 400
        
        # Verify the ride belongs to the current passenger
        ride = Ride.query.get(ride_id)
        if not ride or ride.passenger_id != current_user.id:
            return jsonify({'error': 'Ride not found'}), 404
        
        # Check if feedback already exists
        feedback = Feedback.query.filter_by(ride_id=ride_id).first()
        
        if feedback:
            # Update existing feedback
            feedback.rating = rating
            feedback.comment = comment
        else:
            # Create new feedback
            feedback = Feedback(
                ride_id=ride_id,
                rating=rating,
                comment=comment,
                feedback_type='Rating'
            )
            db.session.add(feedback)
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Rating submitted successfully'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@api.route('/unread-feedback-count')
@admin_required
def get_unread_feedback_count():
    """Get count of unread feedback items"""
    try:
        # Count feedback that hasn't been resolved
        unread_count = Feedback.query.filter_by(is_resolved=False).count()
        
        return jsonify({
            'count': unread_count
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@api.route('/all-feedback')
@admin_required
def get_all_feedback():
    """Get all feedback with ride and passenger details"""
    try:
        feedbacks = Feedback.query.options(
            db.joinedload(Feedback.ride).joinedload(Ride.passenger),
            db.joinedload(Feedback.ride).joinedload(Ride.driver)
        ).order_by(Feedback.submitted_at.desc()).all()
        
        feedback_data = []
        for fb in feedbacks:
            if fb.ride:
                feedback_data.append({
                    'id': fb.id,
                    'ride_id': fb.ride_id,
                    'passenger_name': fb.ride.passenger.username if fb.ride.passenger else 'N/A',
                    'driver_name': fb.ride.driver.name if fb.ride.driver else 'N/A',
                    'rating': fb.rating,
                    'comment': fb.comment,
                    'feedback_type': fb.feedback_type,
                    'details': fb.details,
                    'is_resolved': fb.is_resolved,
                    'submitted_at': to_eat(fb.submitted_at).strftime('%b %d, %Y at %I:%M %p') if fb.submitted_at else 'N/A',
                    'pickup_address': fb.ride.pickup_address,
                    'dest_address': fb.ride.dest_address
                })
        
        return jsonify(feedback_data)
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@api.route('/analytics-data')
@admin_required
def get_analytics_data():
    """Get analytics data for charts and graphs"""
    try:
        period = request.args.get('period', 'week')
        
        # Calculate date range based on period
        now = datetime.now(timezone.utc)
        
        if period == 'today':
            start_date = now.replace(hour=0, minute=0, second=0, microsecond=0)
            prev_start = start_date - timedelta(days=1)
            prev_end = start_date
        elif period == 'week':
            start_date = now - timedelta(days=7)
            prev_start = start_date - timedelta(days=7)
            prev_end = start_date
        elif period == 'month':
            start_date = now - timedelta(days=30)
            prev_start = start_date - timedelta(days=30)
            prev_end = start_date
        elif period == 'year':
            start_date = now - timedelta(days=365)
            prev_start = start_date - timedelta(days=365)
            prev_end = start_date
        else:
            start_date = now - timedelta(days=7)  # Default to week
            prev_start = start_date - timedelta(days=7)
            prev_end = start_date
        
        # Get current period rides
        rides = Ride.query.filter(Ride.request_time >= start_date).all()
        
        # Get previous period rides for trend calculation
        prev_rides = Ride.query.filter(
            Ride.request_time >= prev_start,
            Ride.request_time < prev_end
        ).all()
        
        # Calculate current period statistics
        completed_rides = len([r for r in rides if r.status == 'Completed'])
        canceled_rides = len([r for r in rides if r.status == 'Canceled'])
        active_rides = len([r for r in rides if r.status in ['Assigned', 'On Trip']])
        total_revenue = sum(float(r.fare) for r in rides if r.status == 'Completed')
        avg_fare = total_revenue / completed_rides if completed_rides > 0 else 0
        
        # Calculate previous period statistics for trends
        prev_completed = len([r for r in prev_rides if r.status == 'Completed'])
        prev_revenue = sum(float(r.fare) for r in prev_rides if r.status == 'Completed')
        
        # Calculate trends
        rides_trend = ((completed_rides - prev_completed) / prev_completed * 100) if prev_completed > 0 else 0
        revenue_trend = ((total_revenue - prev_revenue) / prev_revenue * 100) if prev_revenue > 0 else 0
        
        # Vehicle distribution
        vehicle_dist = {}
        for ride in rides:
            vtype = ride.vehicle_type or 'Unknown'
            vehicle_dist[vtype] = vehicle_dist.get(vtype, 0) + 1
        
        # Payment method distribution
        payment_dist = {}
        for ride in rides:
            if ride.status == 'Completed':
                pmethod = ride.payment_method or 'Cash'
                payment_dist[pmethod] = payment_dist.get(pmethod, 0) + 1
        
        # Daily breakdown for revenue chart
        daily_data = {}
        for ride in rides:
            ride_date = to_eat(ride.request_time).strftime('%b %d')
            if ride_date not in daily_data:
                daily_data[ride_date] = 0
            if ride.status == 'Completed':
                daily_data[ride_date] += float(ride.fare)
        
        # Sort by date and prepare chart data
        sorted_dates = sorted(daily_data.keys())
        revenue_chart_data = {
            'labels': sorted_dates[-7:] if len(sorted_dates) > 7 else sorted_dates,  # Last 7 days
            'data': [round(daily_data[d], 2) for d in (sorted_dates[-7:] if len(sorted_dates) > 7 else sorted_dates)]
        }
        
        # Top performing drivers
        driver_stats = db.session.query(
            Driver.id,
            Driver.name,
            Driver.profile_picture,
            func.count(Ride.id).label('completed_rides'),
            func.avg(Feedback.rating).label('avg_rating')
        ).join(Ride, Ride.driver_id == Driver.id)\
         .outerjoin(Feedback, Feedback.ride_id == Ride.id)\
         .filter(Ride.status == 'Completed')\
         .filter(Ride.request_time >= start_date)\
         .group_by(Driver.id)\
         .order_by(func.count(Ride.id).desc())\
         .limit(5).all()
        
        top_drivers = [{
            'name': d.name,
            'avatar': d.profile_picture or 'static/img/default_avatar.png',
            'completed_rides': d.completed_rides,
            'avg_rating': round(float(d.avg_rating), 1) if d.avg_rating else 0
        } for d in driver_stats]
        
        return jsonify({
            'kpis': {
                'rides_completed': completed_rides,
                'active_rides_now': active_rides,
                'rides_canceled': canceled_rides,
            'total_revenue': round(total_revenue, 2),
                'avg_fare': round(avg_fare, 2),
                'trends': {
                    'rides': round(rides_trend, 1),
                    'revenue': round(revenue_trend, 1)
                }
            },
            'charts': {
                'revenue_over_time': revenue_chart_data,
                'vehicle_distribution': vehicle_dist,
                'payment_method_distribution': payment_dist
            },
            'performance': {
                'top_drivers': top_drivers
            }
        })
        
    except Exception as e:
        current_app.logger.error(f"Analytics error: {str(e)}")
        return jsonify({'error': str(e)}), 500

@api.route('/support-tickets')
@admin_required
def get_support_tickets():
    """Get all support tickets"""
    try:
        from app.models import SupportTicket
        
        tickets = SupportTicket.query.options(
            db.joinedload(SupportTicket.passenger),
            db.joinedload(SupportTicket.ride)
        ).order_by(SupportTicket.created_at.desc()).all()
        
        tickets_data = []
        for ticket in tickets:
            tickets_data.append({
                'id': ticket.id,
                'passenger_name': ticket.passenger.username if ticket.passenger else 'N/A',
                'passenger_phone': ticket.passenger.phone_number if ticket.passenger else 'N/A',
                'feedback_type': ticket.feedback_type,
                'details': ticket.details,
                'status': ticket.status,
                'ride_id': ticket.ride_id,
                'created_at': to_eat(ticket.created_at).strftime('%b %d, %Y at %I:%M %p') if ticket.created_at else 'N/A',
                'admin_response': ticket.admin_response
            })
        
        return jsonify(tickets_data)
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@api.route('/support-tickets/<int:ticket_id>/resolve', methods=['POST'])
@admin_required
def resolve_support_ticket(ticket_id):
    """Mark a support ticket as resolved"""
    try:
        from app.models import SupportTicket
        
        ticket = SupportTicket.query.get(ticket_id)
        if not ticket:
            return jsonify({'error': 'Ticket not found'}), 404
        
        data = request.get_json()
        response_text = data.get('response', '')
        
        ticket.status = 'Resolved'
        ticket.admin_response = response_text
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Ticket resolved successfully'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

# --- Driver Earnings System ---

@api.route('/earnings/calculate', methods=['POST'])
@admin_required
def calculate_driver_earnings():
    """Calculate and create earnings records for completed rides"""
    try:
        data = request.json
        ride_id = data.get('ride_id')
        
        if ride_id:
            # Calculate for specific ride
            ride = Ride.query.get(ride_id)
            if not ride or ride.status != 'Completed':
                return jsonify({'error': 'Ride not found or not completed'}), 404
            
            earnings = _create_earnings_record(ride)
            return jsonify({
                'success': True,
                'message': 'Earnings calculated successfully',
                'earnings_id': earnings.id
            })
        else:
            # Calculate for all unprocessed completed rides
            completed_rides = Ride.query.filter(
                Ride.status == 'Completed',
                ~Ride.id.in_(
                    db.session.query(DriverEarnings.ride_id)
                )
            ).all()
            
            created_count = 0
            for ride in completed_rides:
                _create_earnings_record(ride)
                created_count += 1
            
            db.session.commit()
            
            return jsonify({
                'success': True,
                'message': f'Calculated earnings for {created_count} rides',
                'processed_count': created_count
            })
            
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

def _create_earnings_record(ride):
    """Helper function to create earnings record for a ride"""
    # Get commission rate for vehicle type
    commission = Commission.query.filter_by(
        vehicle_type=ride.vehicle_type,
        is_active=True
    ).order_by(Commission.effective_date.desc()).first()
    
    if not commission:
        # Default commission rates if not set
        default_rates = {'Bajaj': 20.0, 'Car': 25.0}
        commission_rate = Decimal(str(default_rates.get(ride.vehicle_type, 20.0)))
    else:
        commission_rate = commission.commission_rate
    
    # Calculate amounts
    gross_fare = ride.fare
    commission_amount = (gross_fare * commission_rate / 100).quantize(Decimal('0.01'))
    driver_earnings = gross_fare - commission_amount
    
    # Create earnings record
    earnings = DriverEarnings(
        driver_id=ride.driver_id,
        ride_id=ride.id,
        gross_fare=gross_fare,
        commission_rate=commission_rate,
        commission_amount=commission_amount,
        driver_earnings=driver_earnings,
        payment_status='Pending'
    )
    
    db.session.add(earnings)
    db.session.commit()
    return earnings

@api.route('/earnings/drivers')
@admin_required
def get_driver_earnings():
    """Get earnings summary for all drivers"""
    try:
        # Get date range from query params
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        driver_id = request.args.get('driver_id')
        
        query = db.session.query(
            Driver.id,
            Driver.name,
            Driver.phone_number,
            Driver.vehicle_type,
            func.count(DriverEarnings.id).label('total_rides'),
            func.sum(DriverEarnings.gross_fare).label('total_fare'),
            func.sum(DriverEarnings.commission_amount).label('total_commission'),
            func.sum(DriverEarnings.driver_earnings).label('total_earnings'),
            func.avg(DriverEarnings.driver_earnings).label('avg_earnings_per_ride')
        ).join(DriverEarnings, Driver.id == DriverEarnings.driver_id)
        
        if start_date:
            query = query.filter(DriverEarnings.created_at >= start_date)
        if end_date:
            query = query.filter(DriverEarnings.created_at <= end_date)
        if driver_id:
            query = query.filter(Driver.id == driver_id)
        
        results = query.group_by(Driver.id).all()
        
        earnings_data = []
        for result in results:
            earnings_data.append({
                'driver_id': result.id,
                'driver_name': result.name,
                'phone_number': result.phone_number,
                'vehicle_type': result.vehicle_type,
                'total_rides': result.total_rides,
                'total_fare': float(result.total_fare or 0),
                'total_commission': float(result.total_commission or 0),
                'total_earnings': float(result.total_earnings or 0),
                'avg_earnings_per_ride': float(result.avg_earnings_per_ride or 0)
            })
        
        return jsonify({
            'success': True,
            'earnings': earnings_data
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@api.route('/earnings/driver/<int:driver_id>')
@admin_required
def get_driver_earnings_detail(driver_id):
    """Get detailed earnings for a specific driver"""
    try:
        driver = Driver.query.get(driver_id)
        if not driver:
            return jsonify({'error': 'Driver not found'}), 404
        
        # Get date range from query params
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        query = DriverEarnings.query.filter_by(driver_id=driver_id)
        
        if start_date:
            query = query.filter(DriverEarnings.created_at >= start_date)
        if end_date:
            query = query.filter(DriverEarnings.created_at <= end_date)
        
        earnings = query.order_by(DriverEarnings.created_at.desc()).all()
        
        earnings_data = []
        for earning in earnings:
            earnings_data.append({
                'id': earning.id,
                'ride_id': earning.ride_id,
                'gross_fare': float(earning.gross_fare),
                'commission_rate': float(earning.commission_rate),
                'commission_amount': float(earning.commission_amount),
                'driver_earnings': float(earning.driver_earnings),
                'payment_status': earning.payment_status,
                'payment_date': earning.payment_date.isoformat() if earning.payment_date else None,
                'created_at': earning.created_at.isoformat(),
                'ride': {
                    'pickup_address': earning.ride.pickup_address,
                    'dest_address': earning.ride.dest_address,
                    'request_time': earning.ride.request_time.isoformat()
                } if earning.ride else None
            })
        
        # Calculate summary
        total_rides = len(earnings)
        total_fare = sum(float(e.gross_fare) for e in earnings)
        total_commission = sum(float(e.commission_amount) for e in earnings)
        total_earnings = sum(float(e.driver_earnings) for e in earnings)
        
        return jsonify({
            'success': True,
            'driver': {
                'id': driver.id,
                'name': driver.name,
                'phone_number': driver.phone_number,
                'vehicle_type': driver.vehicle_type
            },
            'summary': {
                'total_rides': total_rides,
                'total_fare': total_fare,
                'total_commission': total_commission,
                'total_earnings': total_earnings,
                'avg_earnings_per_ride': total_earnings / total_rides if total_rides > 0 else 0
            },
            'earnings': earnings_data
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@api.route('/earnings/update-payment-status', methods=['POST'])
@admin_required
def update_payment_status():
    """Update payment status for driver earnings"""
    try:
        data = request.json
        earnings_id = data.get('earnings_id')
        payment_status = data.get('payment_status')
        
        if not earnings_id or not payment_status:
            return jsonify({'error': 'Missing required fields'}), 400
        
        if payment_status not in ['Pending', 'Paid', 'Disputed']:
            return jsonify({'error': 'Invalid payment status'}), 400
        
        earnings = DriverEarnings.query.get(earnings_id)
        if not earnings:
            return jsonify({'error': 'Earnings record not found'}), 404
        
        earnings.payment_status = payment_status
        if payment_status == 'Paid':
            earnings.payment_date = datetime.now(timezone.utc)
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Payment status updated successfully'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@api.route('/earnings/export')
@admin_required
def export_driver_earnings():
    """Export driver earnings data to CSV"""
    try:
        import csv
        import io
        from flask import make_response
        
        # Get date range from query params
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        query = db.session.query(
            Driver.id,
            Driver.name,
            Driver.phone_number,
            Driver.vehicle_type,
            func.count(DriverEarnings.id).label('total_rides'),
            func.sum(DriverEarnings.gross_fare).label('total_fare'),
            func.sum(DriverEarnings.commission_amount).label('total_commission'),
            func.sum(DriverEarnings.driver_earnings).label('total_earnings'),
            func.avg(DriverEarnings.driver_earnings).label('avg_earnings_per_ride')
        ).join(DriverEarnings, Driver.id == DriverEarnings.driver_id)
        
        if start_date:
            query = query.filter(DriverEarnings.created_at >= start_date)
        if end_date:
            query = query.filter(DriverEarnings.created_at <= end_date)
        
        results = query.group_by(Driver.id).all()
        
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write header
        writer.writerow([
            'Driver ID', 'Driver Name', 'Phone Number', 'Vehicle Type', 
            'Total Rides', 'Total Fare', 'Total Commission', 'Driver Earnings', 'Avg Earnings per Ride'
        ])
        
        # Write data
        for result in results:
            writer.writerow([
                result.id,
                result.name,
                result.phone_number,
                result.vehicle_type,
                result.total_rides,
                float(result.total_fare or 0),
                float(result.total_commission or 0),
                float(result.total_earnings or 0),
                float(result.avg_earnings_per_ride or 0)
            ])
        
        output.seek(0)
        
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'text/csv'
        response.headers['Content-Disposition'] = f'attachment; filename=driver_earnings_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
        
        return response
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@api.route('/export-report')
@admin_required
def export_report():
    """Export comprehensive analytics report in PDF or Excel format"""
    try:
        format_type = request.args.get('format', 'pdf').lower()
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        # Set date range
        if start_date and end_date:
            start_dt = datetime.strptime(start_date, '%Y-%m-%d').replace(tzinfo=timezone.utc)
            end_dt = datetime.strptime(end_date, '%Y-%m-%d').replace(tzinfo=timezone.utc)
            end_dt = end_dt.replace(hour=23, minute=59, second=59)
        else:
            # Default to last 30 days
            end_dt = datetime.now(timezone.utc)
            start_dt = end_dt - timedelta(days=30)
        
        # Get data
        rides = Ride.query.filter(
            Ride.request_time >= start_dt,
            Ride.request_time <= end_dt
        ).options(
            db.joinedload(Ride.passenger),
            db.joinedload(Ride.driver),
            db.joinedload(Ride.feedback)
        ).all()
        
        drivers = Driver.query.all()
        passengers = Passenger.query.all()
        
        if format_type == 'excel':
            return _export_excel_report(rides, drivers, passengers, start_dt, end_dt)
        else:
            return _export_pdf_report(rides, drivers, passengers, start_dt, end_dt)
            
    except Exception as e:
        current_app.logger.error(f"Export report error: {str(e)}")
        return jsonify({'error': str(e)}), 500

def _export_excel_report(rides, drivers, passengers, start_date, end_date):
    """Export report as Excel file"""
    try:
        wb = openpyxl.Workbook()
        
        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        # Add headers
        headers = ['Metric', 'Value']
        for col, header in enumerate(headers, 1):
            cell = ws_summary.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Add summary data
        summary_data = [
            ['Report Period', f"{start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}"],
            ['Total Rides', len(rides)],
            ['Completed Rides', len([r for r in rides if r.status == 'Completed'])],
            ['Total Revenue', sum(float(r.fare) for r in rides if r.status == 'Completed')],
            ['Total Drivers', len(drivers)],
            ['Total Passengers', len(passengers)],
            ['Average Rating', round(sum(r.feedback.rating for r in rides if r.feedback and r.feedback.rating) / len([r for r in rides if r.feedback and r.feedback.rating]), 2) if any(r.feedback and r.feedback.rating for r in rides) else 0]
        ]
        
        for row, (metric, value) in enumerate(summary_data, 2):
            ws_summary.cell(row=row, column=1, value=metric)
            ws_summary.cell(row=row, column=2, value=value)
        
        # Rides sheet
        ws_rides = wb.create_sheet("Rides")
        ride_headers = ['ID', 'Passenger', 'Driver', 'Pickup', 'Destination', 'Fare', 'Status', 'Date', 'Rating']
        for col, header in enumerate(ride_headers, 1):
            cell = ws_rides.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for row, ride in enumerate(rides, 2):
            ws_rides.cell(row=row, column=1, value=ride.id)
            ws_rides.cell(row=row, column=2, value=ride.passenger.username if ride.passenger else 'N/A')
            ws_rides.cell(row=row, column=3, value=ride.driver.name if ride.driver else 'N/A')
            ws_rides.cell(row=row, column=4, value=ride.pickup_address or 'N/A')
            ws_rides.cell(row=row, column=5, value=ride.dest_address)
            ws_rides.cell(row=row, column=6, value=float(ride.fare))
            ws_rides.cell(row=row, column=7, value=ride.status)
            ws_rides.cell(row=row, column=8, value=to_eat(ride.request_time).strftime('%Y-%m-%d %H:%M'))
            ws_rides.cell(row=row, column=9, value=ride.feedback.rating if ride.feedback else 'N/A')
        
        # Auto-adjust column widths
        for ws in [ws_summary, ws_rides]:
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        from flask import make_response
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename=ride_report_{start_date.strftime("%Y%m%d")}_{end_date.strftime("%Y%m%d")}.xlsx'
        
        return response
        
    except Exception as e:
        current_app.logger.error(f"Excel export error: {str(e)}")
        return jsonify({'error': 'Failed to generate Excel report'}), 500

def _export_pdf_report(rides, drivers, passengers, start_date, end_date):
    """Export report as PDF file"""
    try:
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
        styles = getSampleStyleSheet()
        story = []
        
        # Title
        title = Paragraph("Ride Management System - Analytics Report", styles['Title'])
        story.append(title)
        story.append(Spacer(1, 12))
        
        # Report period
        period_text = f"Report Period: {start_date.strftime('%B %d, %Y')} to {end_date.strftime('%B %d, %Y')}"
        period = Paragraph(period_text, styles['Normal'])
        story.append(period)
        story.append(Spacer(1, 20))
        
        # Summary data
        summary_data = [
            ['Metric', 'Value'],
            ['Total Rides', str(len(rides))],
            ['Completed Rides', str(len([r for r in rides if r.status == 'Completed']))],
            ['Total Revenue', f"${sum(float(r.fare) for r in rides if r.status == 'Completed'):.2f}"],
            ['Total Drivers', str(len(drivers))],
            ['Total Passengers', str(len(passengers))],
            ['Average Rating', f"{round(sum(r.feedback.rating for r in rides if r.feedback and r.feedback.rating) / len([r for r in rides if r.feedback and r.feedback.rating]), 2) if any(r.feedback and r.feedback.rating for r in rides) else 0}/5.0"]
        ]
        
        summary_table = Table(summary_data)
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(summary_table)
        story.append(Spacer(1, 20))
        
        # Recent rides table (limit to 20 most recent)
        recent_rides = sorted(rides, key=lambda x: x.request_time, reverse=True)[:20]
        rides_data = [['ID', 'Passenger', 'Driver', 'Destination', 'Fare', 'Status', 'Date']]
        
        for ride in recent_rides:
            rides_data.append([
                str(ride.id),
                ride.passenger.username if ride.passenger else 'N/A',
                ride.driver.name if ride.driver else 'N/A',
                ride.dest_address[:30] + '...' if len(ride.dest_address) > 30 else ride.dest_address,
                f"${float(ride.fare):.2f}",
                ride.status,
                to_eat(ride.request_time).strftime('%m/%d/%Y')
            ])
        
        rides_table = Table(rides_data)
        rides_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 8)
        ]))
        
        story.append(rides_table)
        
        # Build PDF
        doc.build(story)
        buffer.seek(0)
        
        from flask import make_response
        response = make_response(buffer.getvalue())
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'attachment; filename=ride_report_{start_date.strftime("%Y%m%d")}_{end_date.strftime("%Y%m%d")}.pdf'
        
        return response
        
    except Exception as e:
        current_app.logger.error(f"PDF export error: {str(e)}")
        return jsonify({'error': 'Failed to generate PDF report'}), 500



